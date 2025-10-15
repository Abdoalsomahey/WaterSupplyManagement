# views.py
from rest_framework import viewsets, filters
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django.http import HttpResponse
import openpyxl
from openpyxl.utils import get_column_letter

from api.models import Order, User, Notification
from api.serializers import OrderSerializer, DriverOrderSerializer
from api.permissions import IsAdminOrManager, IsDriver
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync

class OrderViewSet(viewsets.ModelViewSet):
    queryset = Order.objects.all()
    serializer_class = OrderSerializer
    permission_classes = [IsAuthenticated, IsAdminOrManager]

    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["id", "status", "customer__full_name", "driver__username"]
    search_fields = ["customer__full_name", "driver__username"]
    ordering_fields = ["created_at", "completed_at", "status"]
    ordering = ["-created_at", "id"]
    
    @action(detail=False, methods=["get"])
    def export_excel(self, request):
        ids = request.GET.getlist("id")
        queryset = self.filter_queryset(self.get_queryset())
        if ids:
            queryset = queryset.filter(id__in=ids)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Orders"

        headers = [
            "ID", "Customer", "Driver", "Status",
            "Created At", "Completed At", "Filled Amount",
            "Is Late", "Failure Reason", "Proof Image"
        ]
        sheet.append(headers)

        for order in queryset:
            sheet.append([
                order.id,
                order.customer.full_name if order.customer else "-",
                order.driver.username if order.driver else "-",
                order.status,
                order.created_at.strftime("%Y-%m-%d %H:%M") if order.created_at else "",
                order.completed_at.strftime("%Y-%m-%d %H:%M") if order.completed_at else "",
                getattr(order, "filled_amount", "-"),
                order.is_driver_late(minutes=30),
                getattr(order, "failure_reason", "-"),
                order.proof_image.url if order.proof_image else "-",
            ])

        for col in sheet.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[col_letter].width = adjusted_width

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="orders.xlsx"'
        workbook.save(response)
        return response


# from openpyxl.drawing.image import Image as ExcelImage
# import os
# from django.http import HttpResponse
# 
# @action(detail=False, methods=['get'])
# def export_excel_with_images(self, request):
#     queryset = self.filter_queryset(self.get_queryset())
#     workbook = openpyxl.Workbook()
#     sheet = workbook.active
#     sheet.title = "Orders"
#
#     headers = [
#         "ID", "Customer", "Driver", "Status",
#         "Created At", "Delivered At", "Filled Amount",
#         "Is Late", "Problem Reason", "Proof Image"
#     ]
#     sheet.append(headers)
#
#     for order in queryset:
#         row = [
#             order.id,
#             order.customer.full_name if order.customer else "-",
#             order.driver.username if order.driver else "-",
#             order.status,
#             order.created_at.strftime("%Y-%m-%d %H:%M") if order.created_at else "",
#             order.completed_at.strftime("%Y-%m-%d %H:%M") if order.completed_at else "",
#             getattr(order, "filled_amount", "-"),
#             order.is_driver_late(minutes=30),
#             getattr(order, "failure_reason", "-"),
#             ""  # proof image placeholder
#         ]
#         sheet.append(row)
#
#         if order.proof_image and os.path.exists(order.proof_image.path):
#             img = ExcelImage(order.proof_image.path)
#             img.width, img.height = 80, 80
#             sheet.add_image(img, f"J{sheet.max_row}")
#
#     response = HttpResponse(
#         content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
#     )
#     response["Content-Disposition"] = 'attachment; filename="orders_with_images.xlsx"'
#     workbook.save(response)
#     return response


class DriverOrderViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = DriverOrderSerializer
    permission_classes = [IsAuthenticated, IsDriver]
    queryset = Order.objects.none()

    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ["status"]
    search_fields = ["customer__full_name"]
    ordering_fields = ["created_at", "completed_at"]
    ordering = ["-created_at"]

    def get_queryset(self):
        return Order.objects.filter(driver=self.request.user)

    @action(detail=True, methods=["post"])
    def confirm(self, request, pk=None):
        order = self.get_object()
        if order.driver != request.user:
            return Response({"error": "This order is not assigned to you"}, status=403)
        if order.status != "pending":
            return Response({"error": "Order already completed or failed"}, status=400)

        filled_amount = request.data.get("filled_amount")
        proof_image = request.FILES.get("proof_image")

        if not filled_amount or not proof_image:
            return Response({"error": "Both filled amount and proof image are required"}, status=400)

        order.confirm(filled_amount=filled_amount, proof_image=proof_image)
        serializer = self.get_serializer(order)
        return Response({
            "message": "Order completed successfully",
            "completed_at": order.completed_at,
            "order": serializer.data
        }, status=200)

    @action(detail=True, methods=["post"])
    def failed(self, request, pk=None):
        order = self.get_object()
        if order.driver != request.user:
            return Response({"error": "This order is not assigned to you"}, status=403)

        reason = request.data.get("reason")
        if not reason:
            return Response({"error": "Failure reason is required"}, status=400)

        order.mark_failed(reason)

        admins_and_managers = User.objects.filter(role__in=["admin", "manager"])
        notifications = []
        for user in admins_and_managers:
            notification = Notification.objects.create(
                recipient=user,
                message=f"Failure reported in Order #{order.id}: {reason}",
                order=order
            )
            notifications.append(notification)

            channel_layer = get_channel_layer()
            async_to_sync(channel_layer.group_send)(
                f"user_{user.id}",
                {
                    "type": "send_notification",
                    "message": notification.message,
                    "order_id": order.id,
                    "created_at": str(notification.created_at),
                }
            )

        serializer = self.get_serializer(order)
        return Response({
            "message": "Order marked as failed",
            "order": serializer.data
        }, status=200)