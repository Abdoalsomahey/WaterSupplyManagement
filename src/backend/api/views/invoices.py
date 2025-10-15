from io import BytesIO
from django.http import HttpResponse
from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
import openpyxl
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4


from api.models import Invoice, User
from api.serializers import InvoiceSerializer, AccountantInvoiceSerializer, SendToAccountantSerializer
from api.permissions import IsAdminOrManager, IsAccountant


class InvoiceViewSet(viewsets.ModelViewSet):
    queryset = Invoice.objects.select_related("customer", "assigned_to").all().order_by("-created_at")
    serializer_class = InvoiceSerializer
    permission_classes = [IsAuthenticated, IsAdminOrManager]

    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["customer__full_name", "customer__phone", "customer__driver__username"]
    filterset_fields = ["id", "status", "assigned_to"]
    ordering_fields = ["period_start", "total_gallons", "total_trips"]

    def get_queryset(self):
        return super().get_queryset()

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=["post"])
    def send_to_accountant(self, request, pk=None):
        serializer = SendToAccountantSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        accountant_username = serializer.validated_data["accountant_username"]

        try:
            accountant = User.objects.get(username=accountant_username, role="accountant")
        except User.DoesNotExist:
            return Response(
                {"error": "Accountant not found or invalid role."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        invoice = self.get_object()
        invoice.assigned_to = accountant
        invoice.status = "sent"
        invoice.save()

        return Response(
            {"detail": f"Invoice sent to accountant '{accountant_username}' successfully."},
            status=status.HTTP_200_OK,
        )
    @action(detail=False, methods=["get"])
    def export_excel(self, request):
        qs = self.filter_queryset(self.get_queryset())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Invoices"
        headers = [
            "Customer", "Phone", "Period Start", "Period End",
            "Total Trips", "Total Gallons", "Price/Gal",
            "Subtotal", "VAT %", "VAT Amount", "Total",
            "Status", "Assigned To", "Created By", "Created At"
        ]
        ws.append(headers)

        for inv in qs:
            ws.append([
                inv.customer.full_name if inv.customer else "",
                inv.customer.phone if inv.customer else "",
                inv.period_start.strftime("%Y-%m-%d") if inv.period_start else "",
                inv.period_end.strftime("%Y-%m-%d") if inv.period_end else "",
                inv.total_trips,
                float(inv.total_gallons),
                float(inv.price_per_gallon or 0),
                float(inv.subtotal),
                float(inv.vat_percent),
                float(inv.vat_amount),
                float(inv.total),
                inv.status,
                inv.assigned_to.username if inv.assigned_to else "",
                inv.created_by.username if inv.created_by else "",
                inv.created_at.strftime("%Y-%m-%d %H:%M") if inv.created_at else "",
            ])

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    l = len(str(cell.value))
                    if l > max_length:
                        max_length = l
            ws.column_dimensions[col_letter].width = max_length + 2

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        resp = HttpResponse(out.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = 'attachment; filename="invoices_admin_manager.xlsx"'
        return resp


class AccountantInvoiceViewSet(viewsets.ModelViewSet):
    queryset = Invoice.objects.select_related("customer", "assigned_to").all().order_by("-created_at")
    serializer_class = AccountantInvoiceSerializer
    permission_classes = [IsAuthenticated, IsAccountant]

    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    search_fields = ["customer__full_name", "customer__phone", "assigned_to__username"]
    filterset_fields = ["id","status", "period_start", "period_end"]
    ordering_fields = ["period_start", "total_gallons", "total_trips", "created_at"]

    def get_queryset(self):
        user = self.request.user
        return self.queryset.filter(assigned_to=user)


    @action(detail=True, methods=["post"])
    def approve(self, request, pk=None):
        invoice = self.get_object()
        if invoice.status != "sent_to_accountant":
            return Response(
                {"error": "Only invoices sent to accountant can be approved."},
                status=status.HTTP_400_BAD_REQUEST
            )
        invoice.status = "approved"
        invoice.save()
        return Response({"detail": f"Invoice #{invoice.id} approved."}, status=status.HTTP_200_OK)

    @action(detail=True, methods=["post"])
    def mark_paid(self, request, pk=None):
        invoice = self.get_object()
        if invoice.status != "approved":
            return Response(
                {"error": "Only approved invoices can be marked as paid."},
                status=status.HTTP_400_BAD_REQUEST
            )
        invoice.status = "paid"
        invoice.save()
        return Response({"detail": f"Invoice #{invoice.id} marked as paid."}, status=status.HTTP_200_OK)

    @action(detail=True, methods=["get"])
    def export_excel(self, request, pk=None):
        invoice = self.get_object()
        customer = invoice.customer

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Invoice_{invoice.id}"

        ws.append(["Customer", customer.full_name if customer else ""])
        ws.append(["Phone", customer.phone if customer else ""])
        ws.append(["Period Start", invoice.period_start.strftime("%Y-%m-%d") if invoice.period_start else ""])
        ws.append(["Period End", invoice.period_end.strftime("%Y-%m-%d") if invoice.period_end else ""])
        ws.append(["Total Trips", invoice.total_trips])
        ws.append(["Total Gallons", float(invoice.total_gallons)])
        ws.append(["Price per Gallon", float(invoice.price_per_gallon or 0)])
        ws.append(["Subtotal", float(invoice.subtotal)])
        ws.append(["VAT %", float(invoice.vat_percent)])
        ws.append(["VAT Amount", float(invoice.vat_amount)])
        ws.append(["Total", float(invoice.total)])
        if invoice.notes:
            ws.append([])
            ws.append(["Notes", invoice.notes])

        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    l = len(str(cell.value))
                    if l > max_length:
                        max_length = l
            ws.column_dimensions[col_letter].width = max_length + 2

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        resp = HttpResponse(out.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        resp["Content-Disposition"] = f'attachment; filename="invoice_{invoice.id}.xlsx"'
        return resp

    @action(detail=True, methods=["get"])
    def export_pdf(self, request, pk=None):
        invoice = self.get_object()
        customer = invoice.customer

        buffer = BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4
        y = height - 50

        p.setFont("Helvetica-Bold", 14)
        p.drawString(50, y, f"Invoice #{invoice.id}")
        y -= 25
        p.setFont("Helvetica", 10)
        p.drawString(50, y, f"Customer: {customer.full_name if customer else ''}")
        y -= 15
        p.drawString(50, y, f"Phone: {customer.phone or ''}")
        y -= 20

        # Header row
        p.setFont("Helvetica-Bold", 10)
        p.drawString(50, y, "Period")
        p.drawString(200, y, "Trips")
        p.drawString(300, y, "Gallons")
        p.drawString(400, y, "Price/gal")
        p.drawString(500, y, "Line Total")
        y -= 15

        p.setFont("Helvetica", 10)
        period_str = f"{invoice.period_start.strftime('%Y-%m-%d')} → {invoice.period_end.strftime('%Y-%m-%d')}" if invoice.period_start and invoice.period_end else ""
        p.drawString(50, y, period_str)
        p.drawString(200, y, str(invoice.total_trips))
        p.drawString(300, y, str(invoice.total_gallons))
        p.drawString(400, y, f"{float(invoice.price_per_gallon or 0):.2f}")
        p.drawString(500, y, f"{float(invoice.subtotal):.2f}")
        y -= 30

        p.drawString(400, y, "Subtotal:")
        p.drawString(500, y, f"{float(invoice.subtotal):.2f}")
        y -= 15
        p.drawString(400, y, f"VAT ({invoice.vat_percent}%):")
        p.drawString(500, y, f"{float(invoice.vat_amount):.2f}")
        y -= 15
        p.setFont("Helvetica-Bold", 11)
        p.drawString(400, y, "TOTAL:")
        p.drawString(500, y, f"{float(invoice.total):.2f}")
        y -= 20

        if invoice.notes:
            p.setFont("Helvetica", 9)
            p.drawString(50, y, f"Notes: {invoice.notes}")

        p.showPage()
        p.save()
        buffer.seek(0)
        return HttpResponse(buffer.read(), content_type="application/pdf")
